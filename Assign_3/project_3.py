import sys
from PySide6 import QtCore, QtWidgets, QtGui
import openpyxl 
from win32com import client
import os

#important fonts used
small_font = QtGui.QFont("Times",12)
small_bold = QtGui.QFont("Times",12, QtGui.QFont.Bold)
large_font = QtGui.QFont("Times",20)
large_font = QtGui.QFont("Times",20)
large_bold_font = QtGui.QFont("Times", 20, QtGui.QFont.Bold)  

class Timecard_Win(QtWidgets.QWidget):
    def __init__(self):
        #parent constructor
        super(Timecard_Win,self).__init__()
        self.setWindowTitle("North Santiam Paving Timesheet")
        self.initUI() 

    def initUI(self):
        """This method contains all the visual pieces of the GUI"""

        #main title for the timesheet
        self.title_1 = QtWidgets.QLabel(self)
        self.title_1.setText("North Santiam Paving Company")
        self.title_1.setFont(small_bold)
        self.title_1.setGeometry(200,50,250,50)

        #Sub-title for the timesheet
        self.title_2 = QtWidgets.QLabel(self)
        self.title_2.setText("Daily Time Card")
        self.title_2.setFont(large_bold_font)
        self.title_2.setStyleSheet("font-style: italic;")
        self.title_2.setGeometry(210,100,250,50)

        #name label
        self._name=''
        self.name_label = QtWidgets.QLabel(self)
        self.name_label.setText('Employee Name:')
        self.name_label.setFont(small_bold)
        self.name_label.move(100,200)
        self.name_label.adjustSize()

        #name button
        self.name = QtWidgets.QPushButton(self)
        self.name.setText(f"{self._name}")
        self.name.setFont(small_font)
        self.name.setStyleSheet("background-color: white;")
        self.name.setGeometry(230,200,150,20)
        self.name.clicked.connect(self.name_update)

        #timecard number label
        self.num_label = QtWidgets.QLabel(self)
        self.num_label.setText("No:")
        self.num_label.setFont(small_font)
        self.num_label.move(1000,50)
        self.num_label.adjustSize()

        #timecard number button
        self._num = ''
        self.num = QtWidgets.QPushButton(self)
        self.num.setText(f'{self._num}')
        self.num.setFont(small_font)
        self.num.setStyleSheet("background-color: white;")
        self.num.setGeometry(1040,50,80,20)
        self.num.clicked.connect(self.num_update)

        #timecard Day label
        self.day_label = QtWidgets.QLabel(self)
        self.day_label.setText('Day:')
        self.day_label.setFont(small_font)
        self.day_label.move(1000,75)
        self.day_label.adjustSize()

        #timecard Day button
        self._day=''
        self.day = QtWidgets.QPushButton(self)
        self.day.setText(f"{self._day}")
        self.day.setFont(small_font)
        self.day.setStyleSheet("background-color: white;")
        self.day.setGeometry(1040,75,80,20)
        self.day.clicked.connect(self.day_update)

        #timecard Date label
        self.date_label = QtWidgets.QLabel(self)
        self.date_label.setText('Date:')
        self.date_label.setFont(small_font)
        self.date_label.move(1000,100)
        self.date_label.adjustSize()

        #timecard Date Button
        self._date=''
        self.date = QtWidgets.QPushButton(self)
        self.date.setText(f'{self._date}')
        self.date.setFont(small_font)
        self.date.setStyleSheet("background-color: white;")
        self.date.setGeometry(1040,100,80,20)
        self.date.clicked.connect(self.date_update)

        #upload button
        self.upload = QtWidgets.QPushButton(self)
        self.upload.setText('Upload Timesheet')
        self.upload.setFont(large_bold_font)
        self.upload.setStyleSheet("background-color: red;")
        self.upload.setGeometry(1040,600,300,75)
        self.upload.clicked.connect(self.upload_excel)

        #download button
        self.download = QtWidgets.QPushButton(self)
        self.download.setText('Download Timesheet')
        self.download.setFont(large_bold_font)
        self.download.setStyleSheet("background-color: lime;")
        self.download.setGeometry(1040,700,300,75)
        self.download.clicked.connect(self.download_pdf)

        #Job # label
        self.num_job_label = QtWidgets.QLabel(self)
        self.num_job_label.setText('Job #')
        self.num_job_label.setFont(small_bold)
        self.num_job_label.setAlignment(QtCore.Qt.AlignCenter)
        self.num_job_label.setStyleSheet("background-color: yellow; border: 2px solid black;")
        self.num_job_label.setGeometry(100,300,72,50)

        #phase label
        self.phase_label = QtWidgets.QLabel(self)
        self.phase_label.setText('Phase')
        self.phase_label.setFont(small_bold)
        self.phase_label.setAlignment(QtCore.Qt.AlignCenter)
        self.phase_label.setStyleSheet("background-color: yellow; border: 2px solid black;")
        self.phase_label.setGeometry(170,300,80,50)

        #job location
        self.job_loc = QtWidgets.QLabel(self)
        self.job_loc.setText('Job Location')
        self.job_loc.setFont(small_bold)
        self.job_loc.setAlignment(QtCore.Qt.AlignCenter)
        self.job_loc.setStyleSheet("background-color: yellow; border: 2px solid black;")
        self.job_loc.setGeometry(238,300,120,50)

        #job description
        self.job_des = QtWidgets.QLabel(self)
        self.job_des.setText('Job Description')
        self.job_des.setFont(small_bold)
        self.job_des.setAlignment(QtCore.Qt.AlignCenter)
        self.job_des.setStyleSheet("background-color: yellow; border: 2px solid black;")
        self.job_des.setGeometry(350,300,200,50)

        #blank spots
        self.blank_spot_1 = QtWidgets.QLabel(self)
        self.blank_spot_1.setStyleSheet("background-color: gray;")
        self.blank_spot_1.setGeometry(548,300,60,300)

        #Job # box
        self.job_1_box = QtWidgets.QPushButton(self)
        self.job_2_box = QtWidgets.QPushButton(self)
        self.job_3_box = QtWidgets.QPushButton(self)
        self.job_4_box = QtWidgets.QPushButton(self)
        self.job_5_box = QtWidgets.QPushButton(self)

        #Phase box
        self.job_1_box_phase = QtWidgets.QPushButton(self)
        self.job_2_box_phase = QtWidgets.QPushButton(self)
        self.job_3_box_phase = QtWidgets.QPushButton(self)
        self.job_4_box_phase = QtWidgets.QPushButton(self)
        self.job_5_box_phase = QtWidgets.QPushButton(self)

        #job number initial string
        self.job_1_num = ''
        self.job_2_num = ''
        self.job_3_num = ''
        self.job_4_num = ''
        self.job_5_num = ''

        #job phase initial string
        self.job_1_phase = ''
        self.job_2_phase = ''
        self.job_3_phase = ''
        self.job_4_phase = ''
        self.job_5_phase = ''

        #job location initial string
        self.job_1_loc = ''
        self.job_2_loc = ''
        self.job_3_loc = ''
        self.job_4_loc = ''
        self.job_5_loc = ''

        #job description initial string
        self.job_1_descr = ''
        self.job_2_descr = ''
        self.job_3_descr = ''
        self.job_4_descr = ''
        self.job_5_descr = ''

        #setting the text of job #
        self.job_1_box.setText(self.job_1_num)
        self.job_2_box.setText(self.job_2_num)
        self.job_3_box.setText(self.job_3_num)
        self.job_4_box.setText(self.job_4_num)
        self.job_5_box.setText(self.job_5_num)

        #setting the font of job #
        self.job_1_box.setFont(small_font)
        self.job_2_box.setFont(small_font)
        self.job_3_box.setFont(small_font)
        self.job_4_box.setFont(small_font)
        self.job_5_box.setFont(small_font)

        #settingt the style of the job # buttons
        self.job_1_box.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_2_box.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_3_box.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_4_box.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_5_box.setStyleSheet("background-color: white; border: 2px solid black;")

        #setting the geometry and postion (x,y,l,w)
        self.job_1_box.setGeometry(100,350,72,50)
        self.job_2_box.setGeometry(100,400,72,50)
        self.job_3_box.setGeometry(100,450,72,50)
        self.job_4_box.setGeometry(100,500,72,50)
        self.job_5_box.setGeometry(100,550,72,50)

        #calling the job entry function to get user input
        #must use lambda functions so these methods are automatically called due to None value
        self.job_1_box.clicked.connect(lambda: self.job_entry(1))
        self.job_2_box.clicked.connect(lambda: self.job_entry(2))
        self.job_3_box.clicked.connect(lambda: self.job_entry(3))
        self.job_4_box.clicked.connect(lambda: self.job_entry(4))
        self.job_5_box.clicked.connect(lambda: self.job_entry(5))

        #setting text for the phase buttons
        self.job_1_box_phase.setText(self.job_1_num)
        self.job_2_box_phase.setText(self.job_2_num)
        self.job_3_box_phase.setText(self.job_3_num)
        self.job_4_box_phase.setText(self.job_4_num)
        self.job_5_box_phase.setText(self.job_5_num)

        #Setting the font of the phase buttons
        self.job_1_box_phase.setFont(small_font)
        self.job_2_box_phase.setFont(small_font)
        self.job_3_box_phase.setFont(small_font)
        self.job_4_box_phase.setFont(small_font)
        self.job_5_box_phase.setFont(small_font)

        #setting the phase style for the buttons
        self.job_1_box_phase.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_2_box_phase.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_3_box_phase.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_4_box_phase.setStyleSheet("background-color: white; border: 2px solid black;")
        self.job_5_box_phase.setStyleSheet("background-color: white; border: 2px solid black;")

        #setting the geometry of the phase buttons (x,y,l,w)
        self.job_1_box_phase.setGeometry(170,350,71,50)
        self.job_2_box_phase.setGeometry(170,400,71,50)
        self.job_3_box_phase.setGeometry(170,450,71,50)
        self.job_4_box_phase.setGeometry(170,500,71,50)
        self.job_5_box_phase.setGeometry(170,550,71,50)

        #must use lambda call so method is called up when button is pressed and not automatically
        self.job_1_box_phase.clicked.connect(lambda: self.phase_entry(1))
        self.job_2_box_phase.clicked.connect(lambda: self.phase_entry(2))
        self.job_3_box_phase.clicked.connect(lambda: self.phase_entry(3))
        self.job_4_box_phase.clicked.connect(lambda: self.phase_entry(4))
        self.job_5_box_phase.clicked.connect(lambda: self.phase_entry(5))

    def name_update(self):
        """Mini-GUI that takes User Input for Name of employee"""
        self._name, done1 = QtWidgets.QInputDialog.getText(self, "Name", 
        "Enter Name: ")
        self.name.setText(f"{self._name}")

    def num_update(self):
        """Mini-GUI that taker User Input for Employee number"""
        self._num, done1 = QtWidgets.QInputDialog.getText(self, "No", 
        "Enter Employee Number: ")
        self.num.setText(f"{self._num}")

    def day_update(self):
        """Mini-GUI that takes User Input for entry of day"""
        self._day, done1 = QtWidgets.QInputDialog.getText(self, "Day", 
        "Enter Day: ")
        self.day.setText(f"{self._day}")

    def date_update(self):
        """Mini-GUI that takes User Input for entry of data"""
        self._date, done1 = QtWidgets.QInputDialog.getText(self, "Date", 
        "Enter Date: ")
        self.date.setText(f"{self._date}")


    def job_entry(self, row: int):
        """Method that sets the text of the GUI with the inputted data of job entries and updates stored data attributes. 
           Mini-GUI are used for inputs."""
        if row == 1:
            self.job_1_num, done1 = QtWidgets.QInputDialog.getText(self, "Job #", 
            "Enter Job Number: ")
            self.job_1_box.setText(f"{self.job_1_num}")

        elif row == 2:
            self.job_2_num, done1 = QtWidgets.QInputDialog.getText(self, "Job #", 
            "Enter Job Number: ")
            self.job_2_box.setText(f"{self.job_2_num}")

        elif row == 3:
            self.job_3_num, done1 = QtWidgets.QInputDialog.getText(self, "Job #", 
            "Enter Job Number: ")
            self.job_3_box.setText(f"{self.job_3_num}")

        elif row == 4:
            self.job_4_num, done1 = QtWidgets.QInputDialog.getText(self, "Job #", 
            "Enter Job Number: ")
            self.job_4_box.setText(f"{self.job_4_num}")

        elif row == 5:
            self.job_5_num, done1 = QtWidgets.QInputDialog.getText(self, "Job #", 
            "Enter Job Number: ")
            self.job_5_box.setText(f"{self.job_5_num}")

    def phase_entry(self, row: int):
        """Method that sets the text of the GUI with the inputted data of phase entries and updates stored data attributes. 
           Mini-GUI are used for inputs."""
        if row == 1:
            self.job_1_phase, done1 = QtWidgets.QInputDialog.getText(self, "Phase", 
           "Enter Job Phase: ")
            self.job_1_box_phase.setText(f"{self.job_1_phase}")

        elif row == 2:
            self.job_2_phase, done1 = QtWidgets.QInputDialog.getText(self, "Phase", 
            "Enter Job Phase: ")
            self.job_2_box_phase.setText(f"{self.job_2_phase}")

        elif row == 3:
            self.job_3_phase, done1 = QtWidgets.QInputDialog.getText(self, "Phase", 
            "Enter Job Phase: ")
            self.job_3_box_phase.setText(f"{self.job_3_phase}")

        elif row == 4:
            self.job_4_phase, done1 = QtWidgets.QInputDialog.getText(self, "Phase", 
            "Enter Job Phase: ")
            self.job_4_box_phase.setText(f"{self.job_4_phase}")

        elif row == 5:
            self.job_5_phase, done1 = QtWidgets.QInputDialog.getText(self, "Phase", 
            "Enter Job Phase: ")
            self.job_5_box_phase.setText(f"{self.job_5_phase}")

    def upload_excel(self):
        """Method that uploads all data onto Excel"""

        #This function is still under construction. Try and Exceptions will be implemented here to catch any exceptions.
        #List of current exceptions caught, but fixed the code:
            #PermisssionEror: Permission denied if excel file open
            #ValueError Cannot convert [] to Excel
        self.xlsx_path = 'NSP Timecard.xlsx'
        wb = openpyxl.load_workbook(self.xlsx_path)# Load excel file
        ws = wb.active #Declare the worksheet as active so it can be edited
        ws['C5'] = self._name #saving name
        ws['K2'] = self._num #saving employee number
        ws['K3'] = self._day #saving day
        ws['K4'] = self._date #saving date

        #be useful for properties maybe make a separate class just storing all the properties just in case
        job_num = [self.job_1_num,self.job_2_num,self.job_3_num,self.job_4_num,self.job_5_num]
        phase = [self.job_1_phase,self.job_2_phase,self.job_3_phase,self.job_4_phase,self.job_5_phase]
        jobloc = []
        jobdes = []
        timei = []
        timef = []
        hours = []
        equip = []

        #Just for fun, may get removed in future
        colA = chr(65)
        colB = chr(66)
        colC = chr(67)
        colD = chr(68)
        colI = chr(73)
        colJ = chr(74)
        colK = chr(75)
        colL = chr(76)

        #Loop that stores data in proper cells within Excel sheet.
        for i in range(1,6):
            row = 7 + 2*int(i)
            ws[colA + str(row)] = job_num[i-1] 
            ws[colB + str(row)] = phase[i-1] 
            #ws[colC + str(row)] = jobloc[i-1]  
            #ws[colD + str(row)] = jobdes[i-1] 
            #ws[colI + str(row)] = timei[i-1] 
            #ws[colJ + str(row)] = timef[i-1] 
            #ws[colK + str(row)] = hours[i-1] 
            #ws[colL + str(row)] = equip[i-1] 
        wb.save(self.xlsx_path)

    def download_pdf(self):
        """Method that downloads the timesheet as a PDF"""
        path_dir = os.path.dirname(os.path.abspath(__file__))  #directory of the folder from Python script
        pdf_path =  path_dir + r"\NSP_Timecard.pdf"            #PDF path
        excel_path =  path_dir + r"\NSP Timecard.xlsx"         #Excel sheet path
        excel = client.Dispatch("Excel.Application")           #Opening the Excel application in order to edit Excel sheet through COM
        sheets = excel.Workbooks.Open(excel_path)              #Opens workbook
        work_sheets = sheets.Worksheets[0]                     #Takes the first sheet, since there is only 1 in the Excel sheet created
        work_sheets.ExportAsFixedFormat(0, pdf_path)           #exports the PDF
        sheets.Close(False)                                    #closes Excel File
        excel.Quit()                                           #closes the App
                
if __name__ == "__main__":   
    app = QtWidgets.QApplication([])        #creates the GUI object
    win =  Timecard_Win()                   #creates the GUI window object
    win.setStyleSheet("background-color: rgb(182, 182, 182); color: black") #background of window
    win.resize(800, 600)                 
    win.show()
    sys.exit(app.exec())                   #will exit the app when pressed X on GUI